from Funciones import Conexion

def obtener_ventas_vendedores():
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT id, cliente, telefono, numero_guia, ubicacion, encargado, total, fecha_entrega, entregado
        FROM entregas
        ORDER BY fecha_entrega DESC, id DESC
    """
    cursor.execute(query)
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def obtener_total_por_vendedor():
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT encargado, COUNT(*) as cantidad, SUM(total) as total_vendido
        FROM entregas
        GROUP BY encargado
        ORDER BY total_vendido DESC
    """
    cursor.execute(query)
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def agregar_a_factura(id_entrega, monto_adicional):
    conexion = Conexion.conectar()
    if conexion is None:
        return False

    try:
        cursor = conexion.cursor()
        # Verificar si está entregado
        query_check = "SELECT entregado FROM entregas WHERE id = %s"
        cursor.execute(query_check, (id_entrega,))
        resultado = cursor.fetchone()
        
        if resultado and resultado[0] == 1:
            print(f"No se puede agregar a factura. El pedido {id_entrega} ya fue entregado.")
            return False
        
        query = """
            UPDATE entregas 
            SET total = total + %s 
            WHERE id = %s
        """
        cursor.execute(query, (monto_adicional, id_entrega))
        conexion.commit()
        cursor.close()
        conexion.close()
        return True
    except Exception as e:
        print(f"Error al actualizar factura: {e}")
        return False

def obtener_tiendas():
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT DISTINCT ubicacion
        FROM entregas
        ORDER BY ubicacion
    """
    cursor.execute(query)
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return [row[0] for row in resultados]

def obtener_entregas_por_tienda(tienda):
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT id, cliente, telefono, numero_guia, ubicacion, encargado, total, fecha_entrega, entregado
        FROM entregas
        WHERE ubicacion = %s
        ORDER BY fecha_entrega DESC, id DESC
    """
    cursor.execute(query, (tienda,))
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def obtener_ventas_por_vendedor_tienda(tienda):
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT encargado, COUNT(*) as cantidad, SUM(total) as total_vendido
        FROM entregas
        WHERE ubicacion = %s
        GROUP BY encargado
        ORDER BY total_vendido DESC
    """
    cursor.execute(query, (tienda,))
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def marcar_como_entregado(id_entrega):
    conexion = Conexion.conectar()
    if conexion is None:
        return False

    try:
        cursor = conexion.cursor()
        query = """
            UPDATE entregas 
            SET entregado = 1 
            WHERE id = %s
        """
        cursor.execute(query, (id_entrega,))
        conexion.commit()
        cursor.close()
        conexion.close()
        return True
    except Exception as e:
        print(f"Error al marcar como entregado: {e}")
        return False

def obtener_entregas_por_fecha(fecha):
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT id, cliente, telefono, numero_guia, ubicacion, encargado, total, fecha_entrega, entregado
        FROM entregas
        WHERE DATE(fecha_entrega) = %s
        ORDER BY id DESC
    """
    cursor.execute(query, (fecha,))
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def obtener_vendedores_unicos():
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT DISTINCT encargado
        FROM entregas
        ORDER BY encargado
    """
    cursor.execute(query)
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return [row[0] for row in resultados]

def obtener_ventas_por_vendedor(vendedor):
    conexion = Conexion.conectar()
    if conexion is None:
        return []

    cursor = conexion.cursor()
    query = """
        SELECT id, cliente, telefono, numero_guia, ubicacion, encargado, total, fecha_entrega, entregado
        FROM entregas
        WHERE encargado = %s
        ORDER BY fecha_entrega DESC, id DESC
    """
    cursor.execute(query, (vendedor,))
    resultados = cursor.fetchall()

    cursor.close()
    conexion.close()

    return resultados

def generar_excel_ventas(vendedor=None, tienda=None, separar_por_tiendas=False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    wb = Workbook()
    
    # Estilos comunes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def llenar_hoja_detalle(ws, datos):
        encabezados = ["ID", "Cliente", "Teléfono", "Número Guía", "Ubicación", "Encargado", "Total", "Fecha Entrega", "Entregado"]

        for col, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = encabezado
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_idx, row in enumerate(datos, start=2):
            ws.cell(row=row_idx, column=1).value = row[0]  # ID
            ws.cell(row=row_idx, column=2).value = row[1]  # Cliente
            ws.cell(row=row_idx, column=3).value = row[2]  # Teléfono
            ws.cell(row=row_idx, column=4).value = row[3]  # Número Guía
            ws.cell(row=row_idx, column=5).value = row[4]  # Ubicación
            ws.cell(row=row_idx, column=6).value = row[5]  # Encargado
            ws.cell(row=row_idx, column=7).value = row[6]  # Total
            ws.cell(row=row_idx, column=8).value = row[7]  # Fecha
            ws.cell(row=row_idx, column=9).value = "Sí" if row[8] == 1 else "No"  # Entregado

            # Aplicar bordes y alineación
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if col == 7:  # Columna de Total
                    cell.number_format = '$#,##0.00'

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12

        # Agregar total
        row_total = len(datos) + 2
        ws.cell(row=row_total, column=6).value = "TOTAL:"
        ws.cell(row=row_total, column=6).font = Font(bold=True)
        
        total_sum = sum([v[6] for v in datos])
        total_cell = ws.cell(row=row_total, column=7)
        total_cell.value = total_sum
        total_cell.font = Font(bold=True)
        total_cell.number_format = '$#,##0.00'

    if separar_por_tiendas:
        wb.remove(wb.active) # Eliminar hoja por defecto
        tiendas = obtener_tiendas()
        hojas_creadas = 0
        for nombre_tienda in tiendas:
            if not nombre_tienda: continue
            
            # Limpieza robusta de nombre de hoja para Excel
            invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
            safe_name = str(nombre_tienda)
            for char in invalid_chars:
                safe_name = safe_name.replace(char, '')
            safe_name = safe_name[:31]
            
            ws = wb.create_sheet(title=safe_name)
            
            ventas = obtener_entregas_por_tienda(nombre_tienda)
            
            # 1. Calcular totales por vendedor para ordenar
            totales_vendedor = {}
            ventas_por_vendedor = {}
            
            for v in ventas:
                encargado = v[5] if v[5] else "Sin Asignar"
                monto = float(v[6]) if v[6] else 0.0
                
                totales_vendedor[encargado] = totales_vendedor.get(encargado, 0) + monto
                
                if encargado not in ventas_por_vendedor:
                    ventas_por_vendedor[encargado] = []
                ventas_por_vendedor[encargado].append(v)
            
            # 2. Ordenar vendedores por total descendente
            vendedores_ordenados = sorted(totales_vendedor.keys(), key=lambda k: totales_vendedor[k], reverse=True)
            
            # 3. Escribir Encabezados (Detallado)
            encabezados = ["ID", "Cliente", "Teléfono", "Número Guía", "Ubicación", "Encargado", "Total", "Fecha Entrega", "Entregado"]
            for col, encabezado in enumerate(encabezados, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = encabezado
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 4. Escribir datos agrupados y detallados
            current_row = 2
            for encargado in vendedores_ordenados:
                sus_ventas = ventas_por_vendedor[encargado]
                
                # Escribir cada venta individual
                for row in sus_ventas:
                    ws.cell(row=current_row, column=1).value = row[0]
                    ws.cell(row=current_row, column=2).value = row[1]
                    ws.cell(row=current_row, column=3).value = row[2]
                    ws.cell(row=current_row, column=4).value = row[3]
                    ws.cell(row=current_row, column=5).value = row[4]
                    ws.cell(row=current_row, column=6).value = row[5]
                    ws.cell(row=current_row, column=7).value = row[6]
                    ws.cell(row=current_row, column=8).value = row[7]
                    ws.cell(row=current_row, column=9).value = "Sí" if row[8] == 1 else "No"
                    
                    # Estilos
                    for col in range(1, 10):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if col == 7:
                            cell.number_format = '$#,##0.00'
                    
                    current_row += 1
                
                # Escribir fila de subtotal del vendedor
                ws.cell(row=current_row, column=6).value = f"TOTAL {encargado}:"
                ws.cell(row=current_row, column=6).font = Font(bold=True)
                ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=current_row, column=6).border = border
                
                ws.cell(row=current_row, column=7).value = totales_vendedor[encargado]
                ws.cell(row=current_row, column=7).font = Font(bold=True)
                ws.cell(row=current_row, column=7).number_format = '$#,##0.00'
                ws.cell(row=current_row, column=7).border = border
                
                # Bordes para el resto de la fila (estética)
                for col in range(1, 10):
                    if col not in [6, 7]:
                        ws.cell(row=current_row, column=col).border = border
                
                current_row += 1
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 12
            
            hojas_creadas += 1
        
        if hojas_creadas == 0:
            wb.create_sheet("Sin Datos")
            
        nombre_archivo = f"Ventas_Detalladas_Por_Tiendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        ws = wb.active
        ws.title = "Ventas"
        
        if vendedor:
            ventas = obtener_ventas_por_vendedor(vendedor)
            nombre_archivo = f"Ventas_{vendedor}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif tienda:
            ventas = obtener_entregas_por_tienda(tienda)
            ventas.sort(key=lambda x: x[5] if x[5] else "")
            nombre_archivo = f"Ventas_Tienda_{tienda}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            ventas = obtener_ventas_vendedores()
            nombre_archivo = f"Todas_las_Ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        llenar_hoja_detalle(ws, ventas)

    return wb, nombre_archivo